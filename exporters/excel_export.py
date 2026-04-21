"""
Excel Exporter — generates multi-sheet Excel report with conditional formatting.
"""
import io
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import CategorizedQuestion
from analyzers.frequency import build_frequency_matrix, build_difficulty_distribution
from analyzers.trends import analyze_trends
from analyzers.predictor import predict_topic_distribution


def export_to_excel(
    categorized: list[CategorizedQuestion],
    predictions: list[dict] | None = None,
    sample_questions: dict[str, list[str]] | None = None,
) -> bytes:
    """
    Generate a comprehensive Excel report.

    Returns:
        Excel file contents as bytes
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet 1: Frequency Matrix
        freq = build_frequency_matrix(categorized, group_by="topic")
        if not freq.empty:
            freq.to_excel(writer, sheet_name="Frequency Matrix")
            _format_frequency_sheet(writer, "Frequency Matrix", freq)

        # Sheet 2: Topic Rankings
        if not freq.empty:
            rankings = freq[["Total", "Percentage", "ConsistencyPct", "ProbabilityScore"]].copy()
            rankings.columns = ["Total Questions", "% Share", "Consistency %", "Probability Score"]
            rankings.to_excel(writer, sheet_name="Topic Rankings")
            _format_rankings_sheet(writer, "Topic Rankings", rankings)

        # Sheet 3: Trend Analysis
        trends = analyze_trends(categorized)
        _write_trends_sheet(writer, trends)

        # Sheet 4: Predictions
        if predictions:
            pred_df = pd.DataFrame(predictions)
            pred_df.to_excel(writer, sheet_name="Predictions", index=False)
            _format_predictions_sheet(writer, "Predictions", pred_df)

        # Sheet 5: Sample Questions
        if sample_questions:
            _write_sample_questions_sheet(writer, sample_questions)

        # Sheet 6: Raw Data
        raw_records = []
        for cq in categorized:
            raw_records.append({
                "Year": cq.question.year,
                "Phase": cq.question.phase,
                "Q#": cq.question.question_number,
                "Section": cq.section,
                "Topic": cq.topic,
                "Sub-Topic": cq.sub_topic,
                "Difficulty": cq.difficulty.value,
                "Confidence": round(cq.confidence, 2),
                "Question Text": cq.question.text[:500],
            })
        if raw_records:
            raw_df = pd.DataFrame(raw_records)
            raw_df.to_excel(writer, sheet_name="Raw Data", index=False)

        # Sheet 7: Difficulty Distribution
        diff_dist = build_difficulty_distribution(categorized)
        if not diff_dist.empty:
            diff_dist.to_excel(writer, sheet_name="Difficulty Distribution")

    return output.getvalue()


def _format_frequency_sheet(writer, sheet_name: str, df: pd.DataFrame):
    """Apply heat map formatting to frequency matrix."""
    ws = writer.sheets[sheet_name]

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Heat map colors for data cells (green gradient)
    max_val = df[[c for c in df.columns if c.isdigit()]].max().max() if len(df) > 0 else 1
    max_val = max(max_val, 1)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Apply color scale to year columns
            if isinstance(cell.value, (int, float)) and col > 2:
                intensity = min(cell.value / max_val, 1.0)
                green = int(200 + 55 * (1 - intensity))
                red = int(255 * (1 - intensity * 0.7))
                color = f"{red:02X}{green:02X}{int(200 * (1-intensity)):02X}"
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Auto-width columns
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _format_rankings_sheet(writer, sheet_name: str, df: pd.DataFrame):
    """Format rankings sheet with color-coded probability scores."""
    ws = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _format_predictions_sheet(writer, sheet_name: str, df: pd.DataFrame):
    """Format predictions sheet."""
    ws = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_trends_sheet(writer, trends: dict):
    """Write trend analysis to a sheet."""
    rows = []

    rows.append({"Category": "DIFFICULTY TREND", "Detail": "", "Value": trends.get("difficulty_trend", "")})
    rows.append({"Category": "", "Detail": "", "Value": ""})

    rows.append({"Category": "RISING TOPICS", "Detail": "", "Value": ""})
    for t in trends.get("rising_topics", []):
        rows.append({
            "Category": "",
            "Detail": f"{t['section']} > {t['topic']}",
            "Value": f"Recent avg: {t['recent_avg']}, Change: {t['change_pct']}%",
        })

    rows.append({"Category": "", "Detail": "", "Value": ""})
    rows.append({"Category": "FALLING TOPICS", "Detail": "", "Value": ""})
    for t in trends.get("falling_topics", []):
        rows.append({
            "Category": "",
            "Detail": f"{t['section']} > {t['topic']}",
            "Value": f"Recent avg: {t['recent_avg']}, Change: {t['change_pct']}%",
        })

    rows.append({"Category": "", "Detail": "", "Value": ""})
    rows.append({"Category": "NEW TOPICS (recent only)", "Detail": "", "Value": ""})
    for t in trends.get("new_topics", []):
        rows.append({
            "Category": "",
            "Detail": f"{t['section']} > {t['topic']}",
            "Value": f"Count: {t['count']}",
        })

    rows.append({"Category": "", "Detail": "", "Value": ""})
    rows.append({"Category": "DISAPPEARED TOPICS", "Detail": "", "Value": ""})
    for t in trends.get("disappeared_topics", []):
        rows.append({
            "Category": "",
            "Detail": f"{t['section']} > {t['topic']}",
            "Value": f"Last count: {t['last_count']}",
        })

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Trend Analysis", index=False)


def _write_sample_questions_sheet(writer, sample_questions: dict[str, list[str]]):
    """Write sample questions to a sheet."""
    rows = []
    for key, questions in sample_questions.items():
        rows.append({"Topic": key, "Question": ""})
        for i, q in enumerate(questions, 1):
            rows.append({"Topic": "", "Question": f"Q{i}: {q}"})
        rows.append({"Topic": "", "Question": ""})

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Sample Questions", index=False)

    ws = writer.sheets["Sample Questions"]
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 100
